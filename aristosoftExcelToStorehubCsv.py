import os
from openpyxl import load_workbook
import codecs
from os import listdir
from os.path import isfile, join

class PrepareThongthanCSV():

    def __init__(self):
        print("Hello World")

    def isExcelFile(self, fileName):
        return not fileName.startswith("~$") and not fileName.startswith(".") and fileName.endswith(".xlsx")

    def doPrepareThongthanCSV():
        excelFiles = [f for f in listdir("./input/") if (isfile(join("./input/", f)) and self.isExcelFile(self, f))]
        # print(excelFiles)
        for fileName in excelFiles:
            fileNameNoExt = os.path.splitext(os.path.basename(fileName))[0]
            excelFileName = "./input/"+fileNameNoExt+".xlsx"
            csvFileName = "./result/"+fileNameNoExt+".csv"
            # print("path: %s" % (os.path.abspath(fileName)))
            book = load_workbook(excelFileName)
            count = 0
            print("----------------------------------")
            try:
                csvData = "SKU,Parent Product SKU,Product Name,Category,Price Type,Unit,Price,Cost,Supplier Price,Product Tags,Inventory Type,Track Stock Levels,Barcode,Variant Name 1,Variant Value 1,Variant Name 2,Variant Value 2,Variant Name 3,Variant Value 3,Thongthan_Quantity,Warning Stock Level,Ideal Stock Level,Supplier,Tax Name,Store Credits,Kitchen Printer\n#Required (Must be unique),,Required,Optional,Required (Fixed/Variable/Unit),Required if the price type is Unit. Leave this field empty if the price Type is either Fixed or Variable.,Optional; 0 by default,Optional,Optional(this field can only be modified if 'Fix Supplier Price' is selected for at least one store in Account Settings),Optional; use semicolon ( ; ) to separate multiple tags,Optional(Simple/Composite/Serialized; Simple by default if tracking inventory),Required (0/1),Optional; Must be unique if specified; use semicolon ( ; ) to separate multiple barcodes,,,,,,,Optional; 0 by default if tracking inventory,Optional,Optional,Optional; use semicolon ( ; ) to separate multiple suppliers,Optional,Optional,Optional\n"
                sizes = list(map(lambda cell: cell.value, book.worksheets[0]["D2":"K2"][0]))
                maxRow = book.worksheets[0].max_row-1
                # print("maxRow: %d" % (maxRow))
                for row in book.worksheets[0].iter_rows(min_row=3, max_row=maxRow, min_col=4, max_col=10):
                    i = 0
                    for cell in row:
                        currentRowA = "B"+str(cell.row)
                        currentRowB = "C"+str(cell.row)
                        currentRowL = "M"+str(cell.row)
                        currentRowN = "O"+str(cell.row)
                        currentNo = list(map(lambda cell: cell.value, book.worksheets[0][currentRowA:currentRowA][0]))[0].strip()
                        currentDesc = list(map(lambda cell: cell.value, book.worksheets[0][currentRowB:currentRowB][0]))[0].strip()
                        # print("currentNo: %s, currentDesc: %s" % (currentNo, currentDesc) )
                        currentPrice = str(float(list(map(lambda cell: cell.value, book.worksheets[0][currentRowL:currentRowL][0]))[0])).strip()
                        # print("currentPrice: %s" % (currentPrice) )
                        currentPriceExt = str(float(list(map(lambda cell: cell.value, book.worksheets[0][currentRowN:currentRowN][0]))[0])).strip()
                        # print("currentPriceExt: %s" % (currentPriceExt) )
                        catagories = "Shirts"
                        # if i == 0:
                        #     csvData += currentNo+",,"+currentNo+" "+currentDesc+","+catagories+",Fixed,,"+currentPrice+","+currentPrice+",,,,1,,,,,,,,,,,,,,\n"
                        if cell.value is None:
                            if i < 6:
                                csvData += currentNo+"_"+sizes[i]+",,"+currentNo+"_"+sizes[i]+" "+currentDesc+","+catagories+",Fixed,,"+currentPrice+","+currentPrice+",,,,1,"+currentNo+"_"+sizes[i]+",,,,,,,"+str(0)+",,,,,,\n"
                            else:
                                csvData += currentNo+"_"+sizes[i]+",,"+currentNo+"_"+sizes[i]+" "+currentDesc+","+catagories+",Fixed,,"+currentPriceExt+","+currentPriceExt+",,,,1,"+currentNo+"_"+sizes[i]+",,,,,,,"+str(0)+",,,,,,\n"
                        else:
                            if i < 6:
                                csvData += currentNo+"_"+sizes[i]+",,"+currentNo+"_"+sizes[i]+" "+currentDesc+","+catagories+",Fixed,,"+currentPrice+","+currentPrice+",,,,1,"+currentNo+"_"+sizes[i]+",,,,,,,"+str(cell.value)+",,,,,,\n"
                            else:
                                csvData += currentNo+"_"+sizes[i]+",,"+currentNo+"_"+sizes[i]+" "+currentDesc+","+catagories+",Fixed,,"+currentPriceExt+","+currentPriceExt+",,,,1,"+currentNo+"_"+sizes[i]+",,,,,,,"+str(cell.value)+",,,,,,\n"
                        i += 1
                        count += 1
                if os.path.exists(csvFileName):
                    print('Remove existing %s' % (csvFileName))
                    os.remove(csvFileName)
                print('Writing data to %s total of %d rows' % (fileNameNoExt+".csv", count))
                f = codecs.open(csvFileName, "w","utf-8")
                f.write(csvData)
                f.close()
                print("----------------------------------")
            except Exception as e:
                print(e)

self = PrepareThongthanCSV
self.doPrepareThongthanCSV()